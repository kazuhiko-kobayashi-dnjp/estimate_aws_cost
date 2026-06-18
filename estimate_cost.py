#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Amazon Bedrock / Claude 用の無料トークンカウンター。
"""
from __future__ import annotations
import argparse, json, os, sys
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    import boto3
    from botocore.exceptions import BotoCoreError, ClientError, ProfileNotFound
except ImportError:
    print("Error: boto3 がインストールされていません。\nインストール: pip install boto3", file=sys.stderr)
    sys.exit(1)

DEFAULT_RATES = {
    "anthropic.claude-sonnet-4-20250514-v1:0": {"input": 3.0, "output": 15.0},
    "anthropic.claude-sonnet-4-6": {"input": 3.0, "output": 15.0},
    "anthropic.claude-opus-4-20250514-v1:0": {"input": 5.0, "output": 25.0},
    "anthropic.claude-opus-4-6-v1": {"input": 5.0, "output": 25.0},
    "anthropic.claude-haiku-4-5-20251001-v1:0": {"input": 1.0, "output": 5.0},
    "anthropic.claude-3-7-sonnet-20250219-v1:0": {"input": 3.0, "output": 15.0},
    "anthropic.claude-3-5-sonnet-20241022-v2:0": {"input": 3.0, "output": 15.0},
    "anthropic.claude-3-5-haiku-20241022-v1:0": {"input": 0.8, "output": 4.0},
}

# 会話ログに記録される model 名（Anthropic API 形式）を、
# CountTokens で実際に使える Bedrock ID とレートへ対応づける。
# CountTokens 非対応の新モデルは、同系列の対応 ID でトークン数を近似計測する。
MODEL_ALIASES = {
    "claude-opus-4-8":   {"count_id": "anthropic.claude-opus-4-6-v1",            "input": 5.0,  "output": 25.0},
    "claude-opus-4-6":   {"count_id": "anthropic.claude-opus-4-6-v1",            "input": 5.0,  "output": 25.0},
    "claude-opus-4":     {"count_id": "anthropic.claude-opus-4-20250514-v1:0",   "input": 5.0,  "output": 25.0},
    "claude-sonnet-4-6": {"count_id": "anthropic.claude-sonnet-4-6",             "input": 3.0,  "output": 15.0},
    "claude-sonnet-4":   {"count_id": "anthropic.claude-sonnet-4-20250514-v1:0", "input": 3.0,  "output": 15.0},
    "claude-haiku-4-5":  {"count_id": "anthropic.claude-haiku-4-5-20251001-v1:0","input": 1.0,  "output": 5.0},
    "claude-fable-5":    {"count_id": "anthropic.claude-opus-4-6-v1",            "input": 10.0, "output": 50.0},
}

DEFAULT_PROFILE = "temp-mfa"
WORKSPACE_ROOT = Path("/home/kazuhiko-kobayashi-dnjp/workspace")



def resolve_model_for_log(log_model: str) -> Tuple[str, float | None, float | None]:
    """ログの model 名を (CountTokens 用 Bedrock ID, 入力レート, 出力レート) に解決する。

    完全一致 → 前方一致（バージョン差を吸収）の順で MODEL_ALIASES を引く。
    未知なら入力をそのまま count_id とし、レートは None（後段で DEFAULT_RATES を試す）。
    """
    if log_model in MODEL_ALIASES:
        a = MODEL_ALIASES[log_model]
        return a["count_id"], a["input"], a["output"]
    for prefix, a in MODEL_ALIASES.items():
        if log_model.startswith(prefix):
            return a["count_id"], a["input"], a["output"]
    # <synthetic> 等の仮想モデル名はWorkflow内部用でAPIコストなし
    if log_model.startswith("<"):
        return "<synthetic>", 0.0, 0.0
    # それ以外の未知モデル名は計測用IDに使えないので未解決(None)扱い
    if not log_model.startswith("claude") and not log_model.startswith("anthropic."):
        return None, None, None
    return log_model, None, None


def to_content_blocks(text: str) -> List[Dict[str, str]]:
    return [{"text": text}]


def build_converse_input_from_text(text: str, system: str | None = None) -> Dict[str, Any]:
    body: Dict[str, Any] = {"messages": [{"role": "user", "content": to_content_blocks(text)}]}
    if system:
        body["system"] = [{"text": system}]
    return body

ALLOWED_TEXT_BLOCK_TYPES = {"text"}


def sanitize_content_blocks(content: Any) -> List[Dict[str, str]]:
    if isinstance(content, str):
        s = content.strip()
        return [{"text": s}] if s else []
    blocks: List[Dict[str, str]] = []
    if isinstance(content, dict):
        if "text" in content and isinstance(content["text"], str):
            s = content["text"].strip()
            if s:
                blocks.append({"text": s})
        return blocks
    if isinstance(content, list):
        for item in content:
            if isinstance(item, str):
                s = item.strip()
                if s:
                    blocks.append({"text": s})
                continue
            if not isinstance(item, dict):
                continue
            item_type = item.get("type")
            if item_type in ALLOWED_TEXT_BLOCK_TYPES and isinstance(item.get("text"), str):
                s = item["text"].strip()
                if s:
                    blocks.append({"text": s})
            elif item_type is None and isinstance(item.get("text"), str):
                s = item["text"].strip()
                if s:
                    blocks.append({"text": s})
        return blocks
    return blocks

VALID_ROLES = {"user", "assistant", "system"}


def build_converse_input_from_conversation_json(raw: Any, system: str | None = None) -> Dict[str, Any]:
    if isinstance(raw, dict) and "messages" in raw:
        messages_raw = raw.get("messages", [])
        system_raw = raw.get("system")
    elif isinstance(raw, list):
        messages_raw = raw
        system_raw = None
    else:
        raise ValueError("conversation JSON は list か {messages: [...]} 形式で指定してください。")
    messages: List[Dict[str, Any]] = []
    for m in messages_raw:
        if not isinstance(m, dict):
            continue
        role = m.get("role")
        if role not in VALID_ROLES or role == "system":
            continue
        blocks = sanitize_content_blocks(m.get("content"))
        if blocks:
            messages.append({"role": role, "content": blocks})
    body: Dict[str, Any] = {"messages": messages}
    resolved_system = system
    if resolved_system is None and system_raw is not None:
        if isinstance(system_raw, str):
            resolved_system = system_raw
        elif isinstance(system_raw, list):
            texts = [b.get("text", "") for b in sanitize_content_blocks(system_raw) if b.get("text")]
            if texts:
                resolved_system = "\n\n".join(texts)
    if resolved_system:
        body["system"] = [{"text": resolved_system}]
    return body


def flatten_message_content(content: Any) -> str:
    """会話 1 メッセージの content を、トークン計測用に 1 本のテキストへ簡約する。

    text / thinking はそのまま、tool_use は input JSON を、tool_result は
    内部の text ブロック（または文字列）を連結する。画像など非テキストは無視。
    """
    if isinstance(content, str):
        return content.strip()
    parts: List[str] = []
    if isinstance(content, list):
        for b in content:
            if isinstance(b, str):
                if b.strip():
                    parts.append(b)
                continue
            if not isinstance(b, dict):
                continue
            t = b.get("type")
            if t == "text" and isinstance(b.get("text"), str):
                parts.append(b["text"])
            elif t == "thinking" and isinstance(b.get("thinking"), str):
                parts.append(b["thinking"])
            elif t == "tool_use":
                parts.append(json.dumps(b.get("input", {}), ensure_ascii=False))
            elif t == "tool_result":
                c = b.get("content")
                if isinstance(c, str):
                    parts.append(c)
                elif isinstance(c, list):
                    for x in c:
                        if isinstance(x, dict) and isinstance(x.get("text"), str):
                            parts.append(x["text"])
                        elif isinstance(x, str):
                            parts.append(x)
    return "\n".join(p for p in parts if p and p.strip()).strip()


def slug_for_cwd(cwd: Path) -> str:
    """カレントディレクトリを Claude Code のプロジェクトログ用スラッグに変換する。

    Claude Code は英数字以外（/ _ . など）をすべて '-' に置換する。
    例: /home/u/workspace/eval_bev_soft -> -home-u-workspace-eval-bev-soft
    """
    import re
    return re.sub(r"[^a-zA-Z0-9]", "-", str(cwd.resolve()))


def list_transcripts() -> List[Path]:
    """カレントプロジェクトの会話ログ (.jsonl) を mtime 降順で返す（[0] が最新）。"""
    base = Path.home() / ".claude" / "projects" / slug_for_cwd(Path.cwd())
    if not base.is_dir():
        return []
    return sorted(base.glob("*.jsonl"), key=lambda p: p.stat().st_mtime, reverse=True)


def find_latest_transcript() -> Path | None:
    """カレントプロジェクトの最新の会話ログ (.jsonl) を返す。見つからなければ None。"""
    logs = list_transcripts()
    return logs[0] if logs else None


def first_user_utterance(path: Path, limit: int = 44) -> str:
    """会話ログ先頭の「本物の」user 発言を要約用に short text で返す。

    スラッシュコマンド・IDE 通知・コマンド出力など、タグだけで中身が空の
    メッセージはスキップし、タグ除去後に文字が残る最初の発話を採用する。
    """
    import re
    try:
        with path.open(encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    o = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if not isinstance(o, dict) or o.get("type") != "user":
                    continue
                msg = o.get("message")
                if not isinstance(msg, dict):
                    continue
                text = flatten_message_content(msg.get("content"))
                text = re.sub(r"<[^>]+>.*?</[^>]+>", " ", text, flags=re.DOTALL)  # 対タグ除去
                text = re.sub(r"<[^>]+>", " ", text)                              # 残った単独タグ
                text = re.sub(r"(?im)^Caveat:.*$", " ", text)                     # コマンド注意書き
                text = " ".join(text.split())
                if text:  # タグ除去後に中身が残る最初の発話＝本物の話題
                    return text[:limit] + ("…" if len(text) > limit else "")
    except OSError:
        pass
    return "(発話なし)"


def resolve_transcript_arg(arg: str) -> Path | None:
    """--transcript の引数を Path に解決する。

    "@latest" → 最新、数字 → 一覧の N 番目、それ以外 → パスとして扱う。
    """
    if arg == "@latest":
        return find_latest_transcript()
    if arg.isdigit():
        logs = list_transcripts()
        idx = int(arg)
        return logs[idx] if 0 <= idx < len(logs) else None
    p = Path(arg)
    return p if p.exists() else None


def print_transcript_list() -> int:
    """カレントプロジェクトの会話ログ一覧を表示する。"""
    import datetime
    logs = list_transcripts()
    if not logs:
        print("カレントプロジェクトの会話ログが見つかりません。", file=sys.stderr)
        return 2
    print(f"\n=== 会話ログ一覧（{len(logs)} 件 / [0] が最新） ===")
    print("-" * 78)
    for i, p in enumerate(logs):
        st = p.stat()
        dt = datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
        kb = st.st_size / 1024
        topic = first_user_utterance(p)
        tag = "  ← 最新" if i == 0 else ""
        print(f"[{i}] {dt}  {kb:7.0f}KB  {topic}{tag}")
    print("-" * 78)
    print("指定例: python3 estimate_cost.py --transcript 0   （番号 or パスで指定）")
    return 0


def build_converse_input_from_transcript(path: Path, system: str | None = None) -> Tuple[Dict[str, Any], int, bool]:
    """Claude Code の JSONL 会話ログを Converse 形式へ変換する。

    戻り値は (converse_input, メッセージ数, 末尾ダミー user を追加したか) のタプル。
    Converse の制約に合わせ、連続する同一 role はマージし、先頭の assistant は除去、
    末尾が assistant の場合はダミー user を 1 件追加する（count_tokens が user 終端を要求するため）。
    """
    raw_msgs: List[Dict[str, str]] = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                o = json.loads(line)
            except json.JSONDecodeError:
                continue
            if not isinstance(o, dict) or o.get("type") not in ("user", "assistant"):
                continue
            msg = o.get("message")
            if not isinstance(msg, dict):
                continue
            role = msg.get("role")
            if role not in ("user", "assistant"):
                continue
            text = flatten_message_content(msg.get("content"))
            if text:
                raw_msgs.append({"role": role, "text": text})
    # 連続する同一 role をマージ
    merged: List[Dict[str, str]] = []
    for m in raw_msgs:
        if merged and merged[-1]["role"] == m["role"]:
            merged[-1]["text"] += "\n" + m["text"]
        else:
            merged.append(dict(m))
    # 先頭の assistant は除去（会話は user で始まる必要がある）
    while merged and merged[0]["role"] == "assistant":
        merged.pop(0)
    # 末尾が assistant なら、ダミー user を追加（count_tokens は user 終端を要求）
    appended_dummy = False
    if merged and merged[-1]["role"] == "assistant":
        merged.append({"role": "user", "text": "."})
        appended_dummy = True
    messages = [{"role": m["role"], "content": [{"text": m["text"]}]} for m in merged]
    body: Dict[str, Any] = {"messages": messages}
    if system:
        body["system"] = [{"text": system}]
    return body, len(messages), appended_dummy


def split_io_texts_from_transcript(path: Path) -> Tuple[str, str, int, int]:
    """会話ログを「入力側」「出力側」のテキストに分離する。

    入力側 = user ロール（あなたの発言・ツール結果）、
    出力側 = assistant ロール（AI が既に生成した返答・思考・ツール呼び出し）。
    戻り値は (入力テキスト, 出力テキスト, 入力メッセージ数, 出力メッセージ数)。
    """
    input_parts: List[str] = []
    output_parts: List[str] = []
    n_in = n_out = 0
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                o = json.loads(line)
            except json.JSONDecodeError:
                continue
            if not isinstance(o, dict) or o.get("type") not in ("user", "assistant"):
                continue
            msg = o.get("message")
            if not isinstance(msg, dict):
                continue
            role = msg.get("role")
            text = flatten_message_content(msg.get("content"))
            if not text:
                continue
            if role == "user":
                input_parts.append(text)
                n_in += 1
            elif role == "assistant":
                output_parts.append(text)
                n_out += 1
    return "\n".join(input_parts), "\n".join(output_parts), n_in, n_out


def _cache_write_tokens_5m(u: Dict[str, Any]) -> int:
    """usage dict から 5分TTL キャッシュ書込トークン数を返す。"""
    cc = u.get("cache_creation")
    if isinstance(cc, dict):
        return int(cc.get("ephemeral_5m_input_tokens", 0))
    # 旧形式: cache_creation サブフィールドなし → 全量を 5m 扱い
    return int(u.get("cache_creation_input_tokens", 0))


def _cache_write_tokens_1h(u: Dict[str, Any]) -> int:
    """usage dict から 1時間TTL キャッシュ書込トークン数を返す。"""
    cc = u.get("cache_creation")
    if isinstance(cc, dict):
        return int(cc.get("ephemeral_1h_input_tokens", 0))
    return 0


def calc_turn_cost(u: Dict[str, Any], in_rate: float, out_rate: float) -> float:
    """1ターン分の usage dict からコストを計算して返す（USD）。
    5分TTL write = base×1.25、1時間TTL write = base×2.0、read = base×0.1。
    """
    w5m = _cache_write_tokens_5m(u)
    w1h = _cache_write_tokens_1h(u)
    return (
        usd_for_tokens(u.get("input_tokens", 0), in_rate)
        + usd_for_tokens(u.get("output_tokens", 0), out_rate)
        + usd_for_tokens(w5m, in_rate * 1.25)
        + usd_for_tokens(w1h, in_rate * 2.0)
        + usd_for_tokens(u.get("cache_read_input_tokens", 0), in_rate * 0.1)
    )


import datetime as _dt


def _empty_grand() -> Dict[str, Any]:
    return {
        "input_tokens": 0, "output_tokens": 0,
        "cache_creation_input_tokens": 0,
        "cache_creation_5m_tokens": 0, "cache_creation_1h_tokens": 0,
        "cache_read_input_tokens": 0,
        "n_in": 0, "n_out": 0, "cost_usd": 0.0, "unknown_model_turns": 0,
    }


def _iter_turns_from_file(path: Path):
    """JSONL ファイルをターン単位でイテレートする。

    assistant ターンごとに (local_date, usage_dict, model_str) を yield する。
    timestamp がなければ local_date=None。user ターンは n_in カウント用に
    (None, None, None) を yield する（呼び出し側で role を区別しないため
    type="user" のみ (local_date, "user", None) を yield）。
    """
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                o = json.loads(line)
            except json.JSONDecodeError:
                continue
            if not isinstance(o, dict):
                continue
            ts = o.get("timestamp")
            local_date = None
            if isinstance(ts, str):
                try:
                    local_date = _dt.datetime.fromisoformat(
                        ts.replace("Z", "+00:00")
                    ).astimezone().date()
                except ValueError:
                    pass
            turn_type = o.get("type")
            if turn_type == "user":
                msg = o.get("message")
                if isinstance(msg, dict) and msg.get("role") == "user":
                    yield (local_date, "user", None)
            elif turn_type == "assistant":
                msg = o.get("message")
                if not isinstance(msg, dict):
                    continue
                u = msg.get("usage")
                if not isinstance(u, dict):
                    continue
                yield (local_date, msg, u)


def _accumulate_turn(totals: Dict[str, Any], u: Dict[str, Any], model: str) -> None:
    """1ターン分の usage を totals に加算する（コスト計算込み）。"""
    totals["n_out"] += 1
    for k in ("input_tokens", "output_tokens",
              "cache_creation_input_tokens", "cache_read_input_tokens"):
        totals[k] += u.get(k, 0)
    totals["cache_creation_5m_tokens"] += _cache_write_tokens_5m(u)
    totals["cache_creation_1h_tokens"] += _cache_write_tokens_1h(u)
    _, in_rate, out_rate = resolve_model_for_log(model)
    if in_rate is None:
        r = DEFAULT_RATES.get(model)
        if r:
            in_rate, out_rate = r["input"], r["output"]
    if in_rate is not None and out_rate is not None:
        totals["cost_usd"] += calc_turn_cost(u, in_rate, out_rate)
    else:
        totals["unknown_model_turns"] += 1


def count_subagents(path: Path) -> int:
    """セッションの subagents/ ディレクトリにある agent-*.jsonl の件数を返す。"""
    sub_dir = path.parent / path.stem / "subagents"
    if not sub_dir.is_dir():
        return 0
    return sum(1 for _ in sub_dir.glob("agent-*.jsonl"))


# Workflow のキャッシュウォームアップ呼び出し（JSONL 未記録）が顕在化するしきい値。
# 観測データ: n_sub=152 で実コストが JSONL 集計の約1.87倍になった。
# n_sub=1~4 では乖離なし。データ点が少ないため係数化せず警告のみとする。
SUBAGENT_WARN_THRESHOLD = 10


def read_usage_from_transcript(path: Path) -> Dict[str, Any] | None:
    """会話ログの assistant ターンから usage を直接集計して返す（セッション合計）。

    subagents/ も含めて集計。seen_keys で重複除去。
    CountTokens フォールバック用: usage が 1 件もなければ None を返す。
    """
    totals: Dict[str, Any] = {
        "input_tokens": 0, "output_tokens": 0,
        "cache_creation_input_tokens": 0,
        "cache_creation_5m_tokens": 0, "cache_creation_1h_tokens": 0,
        "cache_read_input_tokens": 0,
        "n_in": 0, "n_out": 0, "cost_usd": 0.0, "unknown_model_turns": 0,
    }
    seen_keys: set = set()
    found = False
    paths = [path]
    sub_dir = path.parent / path.stem / "subagents"
    if sub_dir.is_dir():
        paths += sorted(sub_dir.glob("*.jsonl"))
    for p in paths:
        for item in _iter_turns_from_file(p):
            local_date, msg_or_role, u = item
            if msg_or_role == "user":
                totals["n_in"] += 1
                continue
            if u is None:
                continue
            msg = msg_or_role
            dedup_key = (
                msg.get("id", ""), msg.get("model", ""),
                u.get("input_tokens", 0), u.get("output_tokens", 0),
                u.get("cache_creation_input_tokens", 0), u.get("cache_read_input_tokens", 0),
            )
            if dedup_key in seen_keys:
                continue
            seen_keys.add(dedup_key)
            found = True
            _accumulate_turn(totals, u, msg.get("model") or "")
    return totals if found else None


def read_usage_by_date_from_transcript(
    path: Path, seen_keys: set
) -> Dict[Any, Dict[str, Any]]:
    """セッション（+ subagents/）をターンの実行日付ごとに集計して返す。

    戻り値: {local_date: totals_dict}。date=None は timestamp なしターン用。
    seen_keys はセッション間重複除去のために呼び出し側で管理する。
    """
    import collections
    by_date: Dict[Any, Dict[str, Any]] = collections.defaultdict(_empty_grand)
    paths = [path]
    sub_dir = path.parent / path.stem / "subagents"
    if sub_dir.is_dir():
        paths += sorted(sub_dir.glob("*.jsonl"))
    for p in paths:
        for item in _iter_turns_from_file(p):
            local_date, msg_or_role, u = item
            if msg_or_role == "user":
                by_date[local_date]["n_in"] += 1
                continue
            if u is None:
                continue
            msg = msg_or_role
            dedup_key = (
                msg.get("id", ""), msg.get("model", ""),
                u.get("input_tokens", 0), u.get("output_tokens", 0),
                u.get("cache_creation_input_tokens", 0), u.get("cache_read_input_tokens", 0),
            )
            if dedup_key in seen_keys:
                continue
            seen_keys.add(dedup_key)
            _accumulate_turn(by_date[local_date], u, msg.get("model") or "")
    return dict(by_date)


def detect_model_from_transcript(path: Path) -> str | None:
    """会話ログに記録された最後の assistant の model 名を返す。なければ None。

    セッション内でモデルは固定のはずだが、念のため最後に使われたものを採る。
    """
    found: str | None = None
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                o = json.loads(line)
            except json.JSONDecodeError:
                continue
            if not isinstance(o, dict):
                continue
            msg = o.get("message")
            if isinstance(msg, dict) and isinstance(msg.get("model"), str) and msg["model"]:
                found = msg["model"]
    return found


class PromptTooLong(Exception):
    """CountTokens の 200,000 tokens/リクエスト上限を超えたことを示す。"""


# CountTokens の 1 リクエスト上限（トークン）。これ未満になるよう分割して数える。
COUNT_TOKENS_MAX = 200_000


class BedrockTokenCounter:
    def __init__(self, profile: str | None, region: str):
        try:
            session = boto3.Session(profile_name=profile, region_name=region) if profile else boto3.Session(region_name=region)
        except ProfileNotFound as e:
            raise SystemExit(f"AWS profile が見つかりません: {e}")
        self.client = session.client("bedrock-runtime", region_name=region)

    def count_converse(self, model_id: str, converse_input: Dict[str, Any]) -> int:
        if not hasattr(self.client, "count_tokens"):
            import botocore
            raise SystemExit(
                "Error: お使いの boto3/botocore は CountTokens API に未対応です。\n"
                f"  現在の botocore: {botocore.__version__}（CountTokens には 1.40 以降が必要）\n"
                "  対処: ライブラリを更新してください。\n"
                "    pip install -U 'boto3>=1.40'\n"
                "  ※ OS 同梱版(/usr/lib/python3/dist-packages)を使っている場合は、\n"
                "     仮想環境(venv)を作って pip で新しい boto3 を入れるのが安全です。"
            )
        try:
            resp = self.client.count_tokens(modelId=model_id, input={"converse": converse_input})
            return int(resp["inputTokens"])
        except ClientError as e:
            msg = str(e)
            if "too long" in msg or "maximum" in msg:
                # CountTokens は 1 リクエスト 200,000 tokens 上限。呼び出し側で分割集計する。
                raise PromptTooLong(msg) from e
            if "doesn't support counting tokens" in msg or "model identifier is invalid" in msg:
                supported = ", ".join(sorted(DEFAULT_RATES))
                raise SystemExit(
                    f"Error: model-id '{model_id}' は Bedrock の CountTokens に対応していません。\n"
                    f"  Bedrock 形式のID（例: anthropic.claude-opus-4-6-v1）を --model-id に指定してください。\n"
                    f"  CountTokens 対応の内蔵ID: {supported}"
                ) from e
            if "AccessDenied" in e.response.get("Error", {}).get("Code", "") or "explicit deny" in msg or "MFA" in msg:
                raise SystemExit(
                    "Error: bedrock:CountTokens へのアクセスが拒否されました（AccessDenied）。\n"
                    "  MFA 必須ポリシーにより、MFA 未認証のクレデンシャルが明示的に拒否されている可能性があります。\n"
                    "  対処: MFA 済みプロファイルを --profile で指定するか、AWS_PROFILE に設定してください。\n"
                    "        例) python3 estimate_cost.py --transcript --profile temp-mfa --show-cost\n"
                    "  一時クレデンシャルが期限切れの場合は MFA セッションを取り直してください。"
                ) from e
            raise SystemExit(f"Error: CountTokens 呼び出しに失敗しました: {msg}") from e
        except BotoCoreError as e:
            raise SystemExit(f"Error: AWS 接続に失敗しました: {e}") from e

    def count_text(self, model_id: str, text: str) -> int:
        """テキストのトークン数を数える。200,000 tokens/リクエスト上限を超える場合は
        テキストを再帰的に分割して数え、合計を返す（トークン数は加算可能）。
        """
        if not text:
            return 0
        try:
            return self.count_converse(model_id, build_converse_input_from_text(text))
        except PromptTooLong:
            # 文字単位で半分に割り、それぞれを数えて合算（1 文字まで割れば必ず上限未満になる）
            if len(text) <= 1:
                raise  # 1 文字で上限超過はあり得ない。想定外なので送出
            mid = len(text) // 2
            return self.count_text(model_id, text[:mid]) + self.count_text(model_id, text[mid:])


def read_text_file(path: Path) -> str:
    for enc in ("utf-8", "utf-8-sig", "cp932", "latin-1"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("unknown", b"", 0, 1, f"{path} を読み込めませんでした")


def usd_for_tokens(tokens: int, usd_per_mtok: float) -> float:
    return (tokens / 1_000_000.0) * usd_per_mtok


def resolve_rates(model_id: str, input_rate: float | None, output_rate: float | None) -> Tuple[float | None, float | None]:
    if input_rate is not None or output_rate is not None:
        return input_rate, output_rate
    if model_id in DEFAULT_RATES:
        r = DEFAULT_RATES[model_id]
        return r["input"], r["output"]
    return None, None


def print_transcript_summary(path: Path, model_id: str, in_tokens: int, out_tokens: int,
                             n_in: int, n_out: int,
                             input_rate: float | None, output_rate: float | None,
                             cache_write_tokens: int = 0, cache_read_tokens: int = 0,
                             from_usage: bool = False) -> None:
    """会話ログの実コスト（既存の入力＋既存の出力）を表示する。"""
    JPY = 150.0
    total_tokens = in_tokens + out_tokens + cache_write_tokens + cache_read_tokens
    source_label = "usage 直読み（実請求値）" if from_usage else "CountTokens 推定値"
    print(f"\n=== セッション実コスト（会話ログ集計 / {source_label}） ===")
    print(f"model_id : {model_id}")
    print(f"log      : {path.name}")
    print("-" * 72)
    print(f"入力         ({n_in:>3} msgs): {in_tokens:>9,} tokens")
    print(f"出力         ({n_out:>3} msgs): {out_tokens:>9,} tokens")
    if from_usage:
        print(f"キャッシュ書込             : {cache_write_tokens:>9,} tokens")
        print(f"キャッシュ読取             : {cache_read_tokens:>9,} tokens")
    print("-" * 72)
    print(f"合計トークン               : {total_tokens:>9,} tokens")
    if input_rate is None or output_rate is None:
        print("\nコスト: レート未設定。--input-rate / --output-rate を指定するか、内蔵モデルIDを使用してください。")
        return
    in_cost = usd_for_tokens(in_tokens, input_rate)
    out_cost = usd_for_tokens(out_tokens, output_rate)
    # キャッシュ書込 = 入力レート×1.25、キャッシュ読取 = 入力レート×0.1（Anthropic 標準比率）
    cache_write_cost = usd_for_tokens(cache_write_tokens, input_rate * 1.25) if from_usage else 0.0
    cache_read_cost  = usd_for_tokens(cache_read_tokens,  input_rate * 0.1)  if from_usage else 0.0
    total_cost = in_cost + out_cost + cache_write_cost + cache_read_cost
    print(f"入力コスト                 : ${in_cost:>10.6f}  (rate=${input_rate}/MTok)")
    print(f"出力コスト                 : ${out_cost:>10.6f}  (rate=${output_rate}/MTok)")
    if from_usage:
        print(f"キャッシュ書込コスト        : ${cache_write_cost:>10.6f}  (rate=${input_rate * 1.25:.4f}/MTok)")
        print(f"キャッシュ読取コスト        : ${cache_read_cost:>10.6f}  (rate=${input_rate * 0.1:.4f}/MTok)")
    line = "=" * 72
    print(f"\n{line}")
    print(f"  ★ 総額: ${total_cost:.4f}  (約 {total_cost * JPY:,.1f} 円)")
    print(line)


def print_summary(rows: List[Dict[str, Any]], model_id: str, input_rate: float | None, output_rate: float | None) -> None:
    if not rows:
        print("結果がありません。")
        return
    total_tokens = sum(int(r["tokens"]) for r in rows)
    print("\n=== CountTokens 結果 ===")
    print(f"model_id : {model_id}")
    print("-" * 72)
    for r in rows:
        label = r["label"]
        chars = r.get("chars")
        desc = f"  文字数 {chars:,}" if chars is not None else ""
        print(f"{label}: {r['tokens']:,} tokens{desc}")
    print("-" * 72)
    print(f"合計     : {total_tokens:,} tokens")
    if input_rate is None:
        print("入力概算 : レート未設定。--input-rate / --output-rate を指定するか、内蔵モデルIDを使用してください。")
        return
    in_cost = usd_for_tokens(total_tokens, input_rate)
    print(f"入力概算 : ${in_cost:.6f}  (rate=${input_rate}/MTok)")
    if output_rate is not None:
        for out_tokens in (100, 500, 1000, 5000):
            total_cost = in_cost + usd_for_tokens(out_tokens, output_rate)
            print(f"  出力 {out_tokens:>5,} tokens 想定総額: ${total_cost:.6f}")
    # ── 一目でわかる総額（代表値: 出力1,000トークン想定 / 円換算 150円/USD） ──
    JPY = 150.0
    if output_rate is not None:
        headline = in_cost + usd_for_tokens(1000, output_rate)
        head_label = "総額（入力＋出力1,000tok想定）"
    else:
        headline = in_cost
        head_label = "総額（入力のみ）"
    line = "=" * 72
    print(f"\n{line}")
    print(f"  ★ {head_label}: ${headline:.4f}  (約 {headline * JPY:,.1f} 円)")
    print(line)


def list_all_transcripts() -> List[Path]:
    """~/.claude/projects 以下の全 .jsonl を mtime 降順で返す。"""
    base = Path.home() / ".claude" / "projects"
    if not base.is_dir():
        return []
    return sorted(base.rglob("*.jsonl"), key=lambda p: p.stat().st_mtime, reverse=True)


def first_message_timestamp(path: Path) -> float | None:
    """JSONL 内の最初のメッセージの timestamp をエポック秒で返す。なければ None。"""
    import datetime as _dt
    try:
        with path.open(encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    o = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if not isinstance(o, dict):
                    continue
                ts = o.get("timestamp")
                if isinstance(ts, str):
                    try:
                        return _dt.datetime.fromisoformat(ts.replace("Z", "+00:00")).timestamp()
                    except ValueError:
                        pass
                elif isinstance(ts, (int, float)):
                    return float(ts)
    except OSError:
        pass
    return None


def session_timestamp(path: Path) -> float:
    """セッション開始時刻をエポック秒で返す。JSONL 内の timestamp が取れなければ mtime で代替。"""
    t = first_message_timestamp(path)
    return t if t is not None else path.stat().st_mtime


def list_workspace_transcripts_for_month(year: int, month: int) -> List[Path]:
    """WORKSPACE_ROOT 配下の全プロジェクトについて、指定月の .jsonl を降順で返す。

    日付判定は JSONL 内の最初のメッセージの timestamp（UTC）をローカル時刻に変換して行う。
    subagents/ 配下のファイルは除外する（メイン JSONL 経由で集計するため）。
    """
    import re, calendar, datetime
    base = Path.home() / ".claude" / "projects"
    if not base.is_dir():
        return []
    prefix = re.sub(r"[^a-zA-Z0-9]", "-", str(WORKSPACE_ROOT.resolve()))
    # 月の開始・終了をローカル時刻のエポック秒で定義
    start_ts = datetime.datetime(year, month, 1, 0, 0, 0).timestamp()
    last_day = calendar.monthrange(year, month)[1]
    end_ts = datetime.datetime(year, month, last_day, 23, 59, 59).timestamp()
    results: List[Path] = []
    for proj_dir in base.iterdir():
        if not proj_dir.is_dir():
            continue
        if not proj_dir.name.startswith(prefix):
            continue
        # glob("*.jsonl") のみ（subagents/ は除外）
        for p in proj_dir.glob("*.jsonl"):
            st = session_timestamp(p)
            if start_ts <= st <= end_ts:
                results.append(p)
    return sorted(results, key=session_timestamp, reverse=True)


def _add_usage(grand: Dict[str, Any], u: Dict[str, Any]) -> None:
    for k in grand:
        grand[k] += u[k]


def print_workspace_month_summary(year: int, month: int) -> int:
    """workspace 配下の全プロジェクトについて指定月の総合計を Excel で出力する。"""
    import datetime
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Error: openpyxl がインストールされていません。\nインストール: pip install openpyxl", file=sys.stderr)
        return 1

    JPY = 150.0
    # list_workspace_transcripts_for_month は「セッション開始月」で絞るが、
    # 隣月跨ぎのセッションも取りこぼさないよう前後1か月を含めて取得し、
    # 実際のターン日付が対象月内のものだけを集計する。
    import calendar as _cal
    prev_year, prev_month = (year, month - 1) if month > 1 else (year - 1, 12)
    next_year, next_month = (year, month + 1) if month < 12 else (year + 1, 1)
    candidate_logs: List[Path] = []
    for y, m in ((prev_year, prev_month), (year, month), (next_year, next_month)):
        candidate_logs += list_workspace_transcripts_for_month(y, m)
    # 重複除去（同じPathが複数月にまたがって返ることはないが念のため）
    seen_paths: set = set()
    logs: List[Path] = []
    for p in candidate_logs:
        if p not in seen_paths:
            seen_paths.add(p)
            logs.append(p)

    if not logs:
        print(f"No logs found for {year}-{month:02d} under workspace.", file=sys.stderr)
        return 2

    target_year, target_month = year, month
    month_start = datetime.date(target_year, target_month, 1)
    month_end   = datetime.date(target_year, target_month, _cal.monthrange(target_year, target_month)[1])

    grand = _empty_grand()
    # rows: [(date, session_start_dt, proj, usage_for_that_date, model, in_rate, out_rate)]
    rows = []
    skipped = 0
    global_seen_keys: set = set()  # セッション間の重複除去
    for p in logs:
        by_date = read_usage_by_date_from_transcript(p, global_seen_keys)
        if not by_date:
            skipped += 1
            continue
        proj = p.parent.name[-40:]
        log_model = detect_model_from_transcript(p) or ""
        _, in_rate, out_rate = resolve_model_for_log(log_model)
        if in_rate is None:
            r2 = DEFAULT_RATES.get(log_model)
            if r2:
                in_rate, out_rate = r2["input"], r2["output"]
        session_dt = datetime.datetime.fromtimestamp(session_timestamp(p))
        n_sub = count_subagents(p)
        for d, u in by_date.items():
            if u["n_out"] == 0:
                continue
            # timestamp なしターン(d=None)は session 開始日に帰属させる
            effective_date = d if d is not None else session_dt.date()
            if not (month_start <= effective_date <= month_end):
                continue
            rows.append({
                "date": effective_date,
                "session_dt": session_dt,
                "proj": proj, "usage": u,
                "model": log_model, "in_rate": in_rate, "out_rate": out_rate,
                "n_sub": n_sub,
            })
            _add_usage(grand, u)

    if skipped:
        print(f"* {skipped} logs skipped (no usage field)", file=sys.stderr)

    rows_asc = sorted(rows, key=lambda r: (r["date"], r["session_dt"]))

    # ── データ行を組み立てる ──
    EMPTY4 = ["", "", "", ""]

    def subtotal_vals(row_type: str, label: str, date_str: str, week_str: str,
                      g: Dict[str, Any]) -> List:
        warn = g["unknown_model_turns"] if g["unknown_model_turns"] else ""
        return [row_type, label, date_str, week_str,
                g["input_tokens"], g["output_tokens"],
                g["cache_creation_input_tokens"], g["cache_read_input_tokens"],
                round(g["cost_usd"], 4), round(g["cost_usd"] * JPY),
                warn, g["n_out"]] + EMPTY4

    out_rows: List[tuple] = []  # (row_type, [values])
    day_grand = _empty_grand()
    week_grand = _empty_grand()
    cur_day: datetime.date | None = None
    cur_week: int | None = None

    def flush_day(day: datetime.date) -> None:
        if day_grand["n_out"] > 0:
            out_rows.append(("day_total", subtotal_vals(
                "day_total", f"{day.strftime('%Y-%m-%d')} total",
                day.strftime("%Y-%m-%d"), f"W{day.isocalendar()[1]:02d}", day_grand,
            )))

    def flush_week(wk: int) -> None:
        if week_grand["n_out"] > 0:
            out_rows.append(("week_total", subtotal_vals(
                "week_total", f"week{wk:02d} total", "", f"W{wk:02d}", week_grand,
            )))

    for r in rows_asc:
        u = r["usage"]
        d = r["date"]
        wk = d.isocalendar()[1]

        if cur_week is not None and wk != cur_week:
            flush_day(cur_day)
            day_grand = _empty_grand()
            flush_week(cur_week)
            week_grand = _empty_grand()
            cur_day = None
        elif cur_day is not None and d != cur_day:
            flush_day(cur_day)
            day_grand = _empty_grand()

        cur_day = d
        cur_week = wk

        n_sub = r.get("n_sub", 0)
        warn_sub = f"⚠n_sub={n_sub}" if n_sub >= SUBAGENT_WARN_THRESHOLD else (n_sub if n_sub else "")
        out_rows.append(("session", [
            "session",
            r["session_dt"].strftime("%Y-%m-%d %H:%M"),
            d.strftime("%Y-%m-%d"),
            f"W{wk:02d}",
            u["input_tokens"],
            u["output_tokens"],
            u["cache_creation_input_tokens"],
            u["cache_read_input_tokens"],
            round(u["cost_usd"], 4),
            round(u["cost_usd"] * JPY),
            u["unknown_model_turns"] or "",
            u["n_out"],
            r["proj"],
            r["model"],
            r["in_rate"] if r["in_rate"] is not None else "",
            r["out_rate"] if r["out_rate"] is not None else "",
            warn_sub,
        ]))
        _add_usage(day_grand, u)
        _add_usage(week_grand, u)

    if cur_day is not None:
        flush_day(cur_day)
    if cur_week is not None:
        flush_week(cur_week)

    out_rows.append(("month_total", subtotal_vals(
        "month_total", f"{year}-{month:02d} total", "", "", grand,
    )))

    # ── Excel 書き込み ──
    HEADER = [
        "type", "label", "date", "iso_week",
        "input_tokens", "output_tokens", "cache_write_tokens", "cache_read_tokens",
        "cost_usd", "cost_jpy", "unresolved_turns", "calls", "project",
        "model", "input_rate", "output_rate", "n_sub",
    ]
    COL_WIDTHS = [12, 22, 12, 8, 14, 14, 16, 14, 10, 10, 14, 8, 44, 30, 11, 12, 12]

    BG_HEADER  = "1F4E79"
    BG_SESSION = ["FFFFFF", "F2F2F2"]
    BG_DAY     = "FFF2CC"
    BG_WEEK    = "FCE4D6"
    BG_MONTH   = "C00000"
    FG_LIGHT   = "FFFFFF"

    def fill(hex6: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex6)

    _thin  = Side(style="thin",   color="BFBFBF")
    _thick = Side(style="medium", color="595959")
    BORDER_NORMAL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    BORDER_SUBTOT = Border(left=_thin, right=_thin, top=_thick, bottom=_thick)

    def apply_border(cells: tuple, brd: Border) -> None:
        for c in cells:
            c.border = brd

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True, color=FG_LIGHT, size=10)
        cell.fill = fill(BG_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"
    apply_border(ws[1], BORDER_SUBTOT)

    FMT_INT  = "#,##0"
    FMT_USD  = "0.0000"
    FMT_JPY  = "#,##0"
    FMT_RATE = "0.00"

    session_idx = 0
    for row_type, values in out_rows:
        ws.append(values)
        row_num = ws.max_row
        cells = ws[row_num]

        if row_type == "session":
            bg = BG_SESSION[session_idx % 2]
            session_idx += 1
            for c in cells:
                c.fill = fill(bg)
                c.font = Font(size=10)
            apply_border(cells, BORDER_NORMAL)
        elif row_type == "day_total":
            for c in cells:
                c.fill = fill(BG_DAY)
                c.font = Font(bold=True, size=10)
            apply_border(cells, BORDER_SUBTOT)
        elif row_type == "week_total":
            for c in cells:
                c.fill = fill(BG_WEEK)
                c.font = Font(bold=True, size=10)
            apply_border(cells, BORDER_SUBTOT)
        elif row_type == "month_total":
            for c in cells:
                c.fill = fill(BG_MONTH)
                c.font = Font(bold=True, color=FG_LIGHT, size=11)
            apply_border(cells, BORDER_SUBTOT)

        for ci in (4, 5, 6, 7):   # input/output/cache token counts (0-based)
            cells[ci].number_format = FMT_INT
        cells[8].number_format = FMT_USD   # cost_usd
        cells[9].number_format = FMT_JPY   # cost_jpy
        cells[11].number_format = FMT_INT  # calls
        if row_type == "session":
            cells[14].number_format = FMT_RATE  # input_rate
            cells[15].number_format = FMT_RATE  # output_rate

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out_path = Path.cwd() / f"cost_{year}-{month:02d}.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}", file=sys.stderr)

    # ── コンソール出力 ──
    JPY_STR = 150.0
    W = 130
    HDR = (
        f"{'日時':<18}"
        f"{'入力':>12}"
        f"{'出力':>12}"
        f"{'C書込':>12}"
        f"{'C読取':>14}"
        f"  {'コスト(USD)':>12}"
        f"  {'calls':>6}"
        f"  {'n_sub':>6}"
        f"  {'モデル':<22}"
        f"  プロジェクト"
    )
    print(f"\n=== workspace 月次集計 {year}-{month:02d} ===")
    print("-" * W)
    print(HDR)
    print("-" * W)

    for row_type, values in out_rows:
        if row_type == "session":
            # values: [type, label, date, week, in, out, cw, cr, cost_usd, cost_jpy, warn, calls, proj, model, in_rate, out_rate, n_sub]
            dt_s   = values[1]
            in_t   = values[4]
            out_t  = values[5]
            cw_t   = values[6]
            cr_t   = values[7]
            cost   = values[8]
            calls  = values[11]
            proj   = values[12]
            model  = values[13]
            n_sub_val = values[16] if len(values) > 16 else ""
            print(
                f"{dt_s:<18}"
                f"{in_t:>12,}"
                f"{out_t:>12,}"
                f"{cw_t:>12,}"
                f"{cr_t:>14,}"
                f"  ${cost:>11.4f}"
                f"  {calls:>6}"
                f"  {str(n_sub_val):>6}"
                f"  {model:<22}"
                f"  {proj}"
            )
        elif row_type in ("day_total", "week_total"):
            in_t   = values[4]
            out_t  = values[5]
            cw_t   = values[6]
            cr_t   = values[7]
            cost   = values[8]
            jpy    = values[9]
            calls  = values[11]
            if row_type == "day_total":
                date_str = values[2]   # "YYYY-MM-DD"
                tag = f"[{date_str[5:7]}/{date_str[8:10]} 小計]"
            else:
                week_str = values[3]   # "W24"
                tag = f"[第{week_str[1:]}週 小計]"
            print(
                f"  {tag:<16}"
                f"{in_t:>12,}"
                f"{out_t:>12,}"
                f"{cw_t:>12,}"
                f"{cr_t:>14,}"
                f"  ${cost:>11.4f}"
                f"  {calls:>6}"
                f"  ({jpy:,}円)"
            )
            if row_type == "week_total":
                print("~" * W)
        elif row_type == "month_total":
            label = values[1]
            in_t  = values[4]
            out_t = values[5]
            cw_t  = values[6]
            cr_t  = values[7]
            cost  = values[8]
            jpy   = values[9]
            calls = values[11]
            print("=" * W)
            print(
                f"  {'[' + label + ']':<16}"
                f"{in_t:>12,}"
                f"{out_t:>12,}"
                f"{cw_t:>12,}"
                f"{cr_t:>14,}"
                f"  ${cost:>11.4f}"
                f"  {calls:>6}"
                f"  ({jpy:,}円)"
            )
            print("=" * W)

    return 0


def print_all_summary() -> int:
    """workspace 以下の全プロジェクトの全ログをターン単位レートで集計して表示する。"""
    import datetime
    JPY = 150.0
    logs = list_all_transcripts()
    if not logs:
        print("会話ログが見つかりません。", file=sys.stderr)
        return 2

    grand: Dict[str, Any] = {
        "input_tokens": 0, "output_tokens": 0,
        "cache_creation_input_tokens": 0, "cache_read_input_tokens": 0,
        "n_in": 0, "n_out": 0, "cost_usd": 0.0, "unknown_model_turns": 0,
    }
    rows = []
    skipped = 0
    for p in logs:
        u = read_usage_from_transcript(p)
        if u is None:
            skipped += 1
            continue
        mtime = datetime.datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
        rows.append({"path": p, "mtime": mtime, "usage": u})
        for k in grand:
            grand[k] += u[k]

    print(f"\n=== workspace 全セッション集計（{len(rows)} ログ / ターン単位レート適用） ===")
    if skipped:
        print(f"  ※ usage なしのログ {skipped} 件はスキップ")
    print("-" * 90)
    print(f"{'日時':<17} {'入力':>10} {'出力':>10} {'C書込':>10} {'C読取':>10} {'コスト(USD)':>12}  ログ")
    print("-" * 90)
    for r in rows:
        u = r["usage"]
        warn = f" ⚠未解決{u['unknown_model_turns']}ターン" if u["unknown_model_turns"] else ""
        print(
            f"{r['mtime']:<17}"
            f" {u['input_tokens']:>10,}"
            f" {u['output_tokens']:>10,}"
            f" {u['cache_creation_input_tokens']:>10,}"
            f" {u['cache_read_input_tokens']:>10,}"
            f" ${u['cost_usd']:>11.4f}"
            f"  {r['path'].name[:36]}{warn}"
        )
    print("=" * 90)
    print(
        f"{'合計':<17}"
        f" {grand['input_tokens']:>10,}"
        f" {grand['output_tokens']:>10,}"
        f" {grand['cache_creation_input_tokens']:>10,}"
        f" {grand['cache_read_input_tokens']:>10,}"
        f" ${grand['cost_usd']:>11.4f}"
    )
    total_tok = sum(grand[k] for k in ("input_tokens", "output_tokens",
                                        "cache_creation_input_tokens", "cache_read_input_tokens"))
    print(f"\n  総トークン : {total_tok:,} tokens")
    if grand["unknown_model_turns"]:
        print(f"  ⚠ レート未解決ターン : {grand['unknown_model_turns']} turns（コストに含まれていません）")
    line = "=" * 90
    print(f"\n{line}")
    print(f"  ★ 総額: ${grand['cost_usd']:.4f}  (約 {grand['cost_usd'] * JPY:,.1f} 円)")
    print(line)
    return 0


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Amazon Bedrock CountTokens API で Claude の入力トークン数を無料カウントします。")
    p.add_argument("--file", "-f", action="append", help="カウント対象ファイル。複数指定可。")
    p.add_argument("--text", "-t", help="カウント対象テキスト。")
    p.add_argument("--stdin", action="store_true", help="標準入力から読む。")
    p.add_argument("--conversation-json", help="会話 JSON ファイルを読む。thinking/tool_use は text のみに簡約。")
    p.add_argument("--transcript", nargs="?", const="@latest", metavar="PATH|N",
                   help="Claude Code の会話ログ (.jsonl) を読んでトークン数を測る。"
                        "省略時は最新ログ。番号(--list の N)・パスでも指定可。")
    p.add_argument("--list", action="store_true", help="カレントプロジェクトの会話ログ一覧（番号・日時・話題）を表示する。")
    p.add_argument("--all", action="store_true", help="~/.claude/projects 以下の全ログをターン単位レートで集計して表示する。")
    p.add_argument("--month", nargs="?", const="current", metavar="YYYY-MM",
                   help="workspace 配下の全プロジェクトについて指定月の総合計を表示。省略時は当月。例: --month 2026-05")
    p.add_argument("--system", help="system prompt を追加。")
    p.add_argument("--model-id", "-m", default=None, help="Bedrock modelId。未指定かつ --transcript 時はログから自動検出。例: anthropic.claude-opus-4-6-v1")
    p.add_argument("--profile", default=os.environ.get("AWS_PROFILE", DEFAULT_PROFILE), help=f"AWS profile 名。未指定時は AWS_PROFILE か '{DEFAULT_PROFILE}'。")
    p.add_argument("--region", default=os.environ.get("AWS_REGION", "us-east-1"), help="AWS region。未指定時は AWS_REGION か us-east-1。")
    p.add_argument("--show-cost", action="store_true", help="概算コストを表示。")
    p.add_argument("--input-rate", type=float, help="入力レート (USD / 1M tokens)。")
    p.add_argument("--output-rate", type=float, help="出力レート (USD / 1M tokens)。")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    if args.list:
        return print_transcript_list()
    if args.all:
        return print_all_summary()
    if args.month is not None:
        import datetime
        if args.month == "current":
            now = datetime.date.today()
            return print_workspace_month_summary(now.year, now.month)
        try:
            dt = datetime.datetime.strptime(args.month, "%Y-%m")
            return print_workspace_month_summary(dt.year, dt.month)
        except ValueError:
            print(f"Error: --month の形式が不正です: '{args.month}'。YYYY-MM で指定してください。", file=sys.stderr)
            return 2
    # 入力が一つも指定されなければ、カレントプロジェクトの最新ログを集計する（既定動作）。
    if not args.file and not args.text and not args.stdin and not args.conversation_json and not args.transcript:
        args.transcript = "@latest"
    counter = BedrockTokenCounter(profile=args.profile, region=args.region)
    rows: List[Dict[str, Any]] = []
    # transcript 以外の入力では従来通り Sonnet を既定 model-id とする。
    fallback_model = args.model_id or "anthropic.claude-sonnet-4-20250514-v1:0"
    if args.text:
        body = build_converse_input_from_text(args.text, system=args.system)
        tokens = counter.count_converse(fallback_model, body)
        rows.append({"label": "text", "tokens": tokens, "chars": len(args.text)})
    if args.stdin:
        text = sys.stdin.read()
        body = build_converse_input_from_text(text, system=args.system)
        tokens = counter.count_converse(fallback_model, body)
        rows.append({"label": "stdin", "tokens": tokens, "chars": len(text)})
    if args.file:
        for fp in args.file:
            path = Path(fp)
            if not path.exists():
                print(f"警告: ファイルが見つかりません: {fp}", file=sys.stderr)
                continue
            text = read_text_file(path)
            # 大きいファイルは 200,000 tokens/リクエスト上限に当たるため分割集計する
            full = (args.system + "\n" + text) if args.system else text
            tokens = counter.count_text(fallback_model, full)
            rows.append({"label": str(path), "tokens": tokens, "chars": len(text)})
    if args.conversation_json:
        path = Path(args.conversation_json)
        if not path.exists():
            print(f"Error: 会話 JSON ファイルが見つかりません: {path}", file=sys.stderr)
            return 2
        raw = json.loads(read_text_file(path))
        body = build_converse_input_from_conversation_json(raw, system=args.system)
        tokens = counter.count_converse(fallback_model, body)
        rows.append({"label": f"conversation:{path}", "tokens": tokens})
    if args.transcript:
        path = resolve_transcript_arg(args.transcript)
        if path is None:
            print(f"Error: 会話ログを解決できません: '{args.transcript}'。--list で一覧を確認してください。", file=sys.stderr)
            return 2
        # モデルの決定: --model-id 明示 > ログから自動検出 > Sonnet 既定
        SONNET = "anthropic.claude-sonnet-4-20250514-v1:0"
        log_model = detect_model_from_transcript(path)
        if args.model_id:
            count_id, in_rate, out_rate = args.model_id, None, None
            display_model = args.model_id
        elif log_model:
            count_id, in_rate, out_rate = resolve_model_for_log(log_model)
            if count_id is None:
                count_id, display_model = SONNET, f"{log_model} (計測: {SONNET})"
            else:
                display_model = f"{log_model} (計測: {count_id})" if count_id != log_model else log_model
        else:
            count_id, in_rate, out_rate = SONNET, None, None
            display_model = count_id
        # --input-rate / --output-rate / DEFAULT_RATES の順でレートを上書き解決
        r_in, r_out = resolve_rates(count_id, args.input_rate, args.output_rate)
        in_rate  = args.input_rate  if args.input_rate  is not None else (in_rate  if in_rate  is not None else r_in)
        out_rate = args.output_rate if args.output_rate is not None else (out_rate if out_rate is not None else r_out)

        # 優先1: JSONL の message.usage を直接集計（実請求値）
        usage = read_usage_from_transcript(path)
        if usage is not None:
            n_sub = count_subagents(path)
            print_transcript_summary(
                path, display_model,
                in_tokens=usage["input_tokens"],
                out_tokens=usage["output_tokens"],
                n_in=usage["n_in"],
                n_out=usage["n_out"],
                input_rate=in_rate,
                output_rate=out_rate,
                cache_write_tokens=usage["cache_creation_input_tokens"],
                cache_read_tokens=usage["cache_read_input_tokens"],
                from_usage=True,
            )
            if n_sub >= SUBAGENT_WARN_THRESHOLD:
                print(f"\n⚠ 警告: このセッションは Workflow subagent を {n_sub} 件含みます。")
                print(  "  Workflow の並列 subagent 起動時、キャッシュウォームアップ呼び出しが")
                print(  "  JSONL に記録されず実請求額が上記集計より大幅に高くなる場合があります。")
                print(  "  実績: n_sub=152 で実コスト ≈ JSONL 集計の 1.87 倍。")
                print(  "  正確な確認は AWS Bedrock モデル呼び出しログ（CloudWatch Logs）を参照してください。")
            return 0

        # フォールバック: usage フィールドがない古いログは CountTokens で推定
        in_text, out_text, n_in, n_out = split_io_texts_from_transcript(path)
        if n_in == 0 and n_out == 0:
            print(f"警告: 会話ログにテキストメッセージがありません: {path}", file=sys.stderr)
            return 0
        in_tokens  = counter.count_text(count_id, in_text)
        out_tokens = counter.count_text(count_id, out_text)
        print_transcript_summary(path, display_model, in_tokens, out_tokens, n_in, n_out, in_rate, out_rate)
        return 0
    input_rate, output_rate = (None, None)
    if args.show_cost:
        input_rate, output_rate = resolve_rates(fallback_model, args.input_rate, args.output_rate)
    print_summary(rows, fallback_model, input_rate, output_rate)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
